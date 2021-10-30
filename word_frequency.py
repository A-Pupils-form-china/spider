import os
import openpyxl
from nltk import word_tokenize, pos_tag
from nltk.corpus import wordnet
from nltk.stem import WordNetLemmatizer
import re
import json

data = dict(json.load(open("../整活/data.txt", 'r')))
stopword = open("stopword.txt", 'r').read().split('\n')
total_word_count = {}
each_word = {}
track5 = {}
track5_sorted = {}
wb = openpyxl.Workbook()


def main():
    def get_wordnet_pos(tag):
        if tag.startswith('J'):
            return wordnet.ADJ
        elif tag.startswith('V'):
            return wordnet.VERB
        elif tag.startswith('N'):
            return wordnet.NOUN
        elif tag.startswith('R'):
            return wordnet.ADV
        else:
            return None

    index = 0
    love = wb.create_sheet(index=11, title="love")
    lovey = 1
    for i in data.items():
        if i[0] == 'album':
            continue
        if i[0] == 'Fearless':
            continue
        album_word = {}
        xy = [1, 1]
        wb.create_sheet(index=index, title=i[0])
        print(index)
        sheet = wb.worksheets[index]
        index += 1
        track = 1
        for j in i[1].values():
            with open("txt/%s/%s.txt" % (
                    i[0].replace(' ', ''),
                    j.replace(' ', '') if not j.__contains__('?') else j.replace(' ', '').replace('?', '')), 'r',
                      encoding='utf8') as r:
                sentences = r.read().split('\n')
            word_count_list = []
            for sentence in sentences:
                word_count = {}
                pattern = re.compile(r"[^a-zA-Z']")
                sentence = re.sub(pattern, ' ', sentence)
                tokens = word_tokenize(sentence)
                tagged_sent = pos_tag(tokens)
                wnl = WordNetLemmatizer()
                for tag in tagged_sent:
                    word_pos = get_wordnet_pos(tag[1]) or wordnet.NOUN
                    word = wnl.lemmatize(tag[0].lower(), pos=word_pos)
                    if word == 'wan':
                        word = 'want'
                    if len(word) > 2:
                        word_count[word] = word_count.get(word, 0)
                        word_count[word] += 1
                word_count_list.append(word_count)
            if track == 5:
                track5[j] = {}
            for word_count in word_count_list:
                for word, count in word_count.items():
                    total_word_count[word] = total_word_count.get(word, 0) + count
                    album_word[word] = album_word.get(word, 0) + count
                    if track == 5:
                        track5[j][word] = track5[j].get(word, 0) + count
            if track == 5:
                listed = list(track5[j].items())
                track5_sorted[j] = []
                for temp in range(len(track5[j])):
                    max = 0
                    max_index = 0
                    for x in range(len(listed)):
                        if listed[x][1] > max:
                            max = listed[x][1]
                            max_index = x
                    track5_sorted[j].append(listed[max_index])
                    listed.remove(listed[max_index])
            track += 1
        album_word_listed = list(album_word.items())
        sorted_data = []
        for sort_temp in range(len(album_word_listed)):
            max = 0
            max_index = 0
            for j in range(len(album_word_listed)):
                if album_word_listed[j][1] > max:
                    max = album_word_listed[j][1]
                    max_index = j
            sorted_data.append(album_word_listed[max_index])
            album_word_listed.remove(album_word_listed[max_index])
        temp = 1
        for sort_temp in sorted_data:
            if temp % 91 == 0:
                xy[0] = 1
                xy[1] += 2
            if sort_temp[0] == 'love':
                love.cell(lovey, 1, i[0])
                love.cell(lovey, 2, sort_temp[1])
                lovey += 1
            sheet.cell(xy[0], xy[1], sort_temp[0])
            sheet.cell(xy[0], xy[1] + 1, sort_temp[1])
            xy[0] += 1
            temp += 1
        each_word[i[0]] = sorted_data[:10]
    sorted_data = []
    listed_data = list(total_word_count.items())
    for i in range(len(listed_data)):
        max = 0
        max_index = 0
        for j in range(len(listed_data)):
            if listed_data[j][1] > max:
                max = listed_data[j][1]
                max_index = j
        sorted_data.append(listed_data[max_index])
        listed_data.remove(listed_data[max_index])
    temp = 1
    a = [1, 1]
    wb.create_sheet(index=10, title='总数据')
    sheet = wb.worksheets[10]
    for i in sorted_data:
        if temp % 250 == 0:
            a[0] = 1
            a[1] += 2
        sheet.cell(a[0], a[1], i[0])
        sheet.cell(a[0], a[1] + 1, i[1])
        temp += 1
        a[0] += 1

    def extra():
        origin_index = 12
        count = 0
        for number in track5_sorted.items():
            sheet = wb.create_sheet(index=origin_index + count, title=number[0].replace("?", ""))
            count += 1
            y = 1
            for j in list(number[1]):
                sheet.cell(y, 1, j[0])
                sheet.cell(y, 2, j[1])
                y += 1
        wb.save("data_without_modify.xlsx")

    extra()


main()
